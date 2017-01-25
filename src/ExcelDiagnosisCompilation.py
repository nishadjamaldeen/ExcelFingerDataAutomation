#####################################################################
#Program Name: Excel Diagnosis (for Hand Arthritis Algorithm)       #  
#Author: Nishad Jamaldeen                                           #   
#Location: eTreatMD Downtown                                        #
#Date Created: 18 January                                           #   
#Date Mofified: 25 January                                          #
#Description: Searches through a folder tree for ourput data from   #
#   the MATLAB algorithm, compiling all avaialible information      #
#   into a single file for analysis and presentation                #
#####################################################################


import xlrd, xlwt, os, time
from os import listdir
from xlrd import open_workbook
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment
from xlutils.copy import copy as copyxl
import pandas as pd



newFileWB = Workbook()

def main():

    startTime = time.time()
    print("Processing...")

    cPath = os.path.dirname(os.path.abspath(__file__))
    row = 1
    compiledDataDestination = createExcelSheet('diagnosisCompilation', cPath)
    newFileWB.save(compiledDataDestination)

    
    subdir = get_sd(cPath)

    for subfolder in subdir:

        directory = os.listdir(subfolder)
        for files in directory:

            if files.endswith('_diagnosis.xls') or files.endswith('_diagnosis.xlsx'):
                diagnosis_filePath = str(subfolder) + "\\" + str(files)
				    
                excelOperations(diagnosis_filePath, subfolder, compiledDataDestination, row)
                row = row + 1

    runTime = time.time() - startTime
    print("Process executed in ", round(runTime,3), " seconds\n")

def get_sd(path):
    return filter(os.path.isdir, [os.path.join(path,f) for f in os.listdir(path)])
	
def rev(s):
    return ''.join(reversed(s))

def createExcelSheet(fileName, destinationPath):

    compiledDataDestination = str(destinationPath) + '\\' + str(fileName) + '.xlsx'
    return compiledDataDestination
	
	
def excelOperations(diagnosis_filePath, subfolder, compiledDataDestination, row):
    try:

        wb = xlrd.open_workbook(diagnosis_filePath)
        sheetNamesArray = wb.sheet_names()
        workingSheet = wb.sheet_by_name(sheetNamesArray[0])
		
        data = pd.read_excel(diagnosis_filePath, header=None)
        data.drop([0])
        data = data.iloc[2:10,:]
        dataMatrix = data.as_matrix()
        print (dataMatrix)

 
		
        subfolderName = ''
        for c in reversed(subfolder):
            if (c == '\\'):
                break
            else:
                subfolderName = subfolderName + c
        
        subfolderName = rev(subfolderName)


        savedFile = load_workbook(compiledDataDestination)


        filesheet = savedFile.active

        filesheet.merge_cells('S1:Z1')
        filesheet['S1'].alignment = Alignment(horizontal='center')
        filesheet['S1'] = 'Joint Angles (Threshold: 7' + str(chr(248)) + ')'

        for i in range(83,91):
            cell = str(chr(i)) + '2'
            filesheet[cell] = dataMatrix[i-83][3]

        joints = ['Index DIP', 'Index PIP', 'Middle DIP', 'Middle PIP', 'Ring DIP',
                  'Ring PIP', 'Little DIP', 'Little PIP']

        for i in range(0,15,2):
            if(66+i <91):
                cell1 = str(chr(66+i)) + '1'
                cell2 = str(chr(67+i)) + '1'
                filesheet.merge_cells(str(cell1) + ':' + str(cell2))

                filesheet[cell1].alignment = Alignment(horizontal='center')
                filesheet[cell1] = joints[int(i/2)]
                filesheet[str(chr(66+i)) + '2'] = 'Width'
                filesheet[str(chr(67+i)) + '2'] = 'Threshold'

        filesheet.cell(column=1,row=row+2, value=subfolderName)

        for c in range(2,18,2):
            filesheet.cell(column=c, row =row+2, value=dataMatrix[int((c-3)/2)][1])

        for c in range(3,18,2):
            filesheet.cell(column=c, row = row+2,value=dataMatrix[int((c-3)/2)][2])

        filesheet.merge_cells('AC1:AJ1')
        filesheet['AC1'] = 'Deformity Diagnosis'
        filesheet['AC1'].alignment = Alignment(horizontal='center')
        
        for i in range(0,8):
            filesheet['A' + str(chr(67+i)) + '2'] = joints[i] + ' Diagnosis'

        for c in range(19,27):
            filesheet.cell(column=c, row = row+2, value=dataMatrix[(c-19)][4])

        for c in range(29,37):
            filesheet.cell(column=c, row=row+2, value=dataMatrix[(c-29)][5])
                          

        savedFile.save(compiledDataDestination)
        
    except FileNotFoundError:
        print("That is an invalid path")

main()
    
